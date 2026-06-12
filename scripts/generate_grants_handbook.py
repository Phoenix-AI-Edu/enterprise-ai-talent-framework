# -*- coding: utf-8 -*-
"""
generate_grants_handbook.py — 建立政府補助與核銷實戰手冊 Excel (.xlsx) 檔案
========================================================================
[Confidential - Phoenix AI Internal Asset]
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure UTF-8 output
if sys.platform.startswith("win"):
    sys.stdout.reconfigure(encoding="utf-8")

def create_handbook():
    wb = Workbook()
    
    # Define common styles
    font_family = "Microsoft JhengHei" # 微軟正黑體 for Chinese compatibility
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(fill_type="solid", start_color="1E293B", end_color="1E293B") # Dark slate
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="334155", end_color="334155") # Slate gray
    
    section_font = Font(name=font_family, size=11, bold=True, color="1E293B")
    section_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0") # Light gray-blue
    
    regular_font = Font(name=font_family, size=10)
    bold_regular_font = Font(name=font_family, size=10, bold=True)
    
    accent_font = Font(name=font_family, size=11, bold=True, color="0D9488") # Teal
    accent_fill = PatternFill(fill_type="solid", start_color="F0FDFA", end_color="F0FDFA") # Light teal tint
    
    alert_fill = PatternFill(fill_type="solid", start_color="FEF2F2", end_color="FEF2F2") # Light red tint
    alert_font = Font(name=font_family, size=10, color="991B1B")
    
    border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    thick_bottom = Border(bottom=Side(style="medium", color="1E293B"))
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row_alt_fill = PatternFill(fill_type="solid", start_color="F8FAFC", end_color="F8FAFC")
    
    # ---------------------------------------------------------
    # Sheet 1: 👑 補助素材盤點與套利規畫表
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "👑 補助素材盤點與套利規畫"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title block
    ws1.merge_cells("A1:D1")
    ws1["A1"] = "鳳凰 AI 補助潛在素材盤點與多層次套利規劃表"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 40
    
    # Part A
    ws1.merge_cells("A2:D2")
    ws1["A2"] = "【Part A：未來 12 個月企業 AI 與數位轉型投資預算盤點】"
    ws1["A2"].font = section_font
    ws1["A2"].fill = section_fill
    ws1["A2"].alignment = align_left
    ws1.row_dimensions[2].height = 25
    
    # Part A headers
    headers_a = ["投資項目類別", "項目細節描述", "實際預估投入金額 (NT$)", "對應公司內部預算科目"]
    for col_idx, h in enumerate(headers_a, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws1.row_dimensions[3].height = 25
    
    part_a_rows = [
        ("1. 研發人員薪資", "AI 演算法工程師與系統架構師內部研發工時", "", "50-薪資支出 (研發部門)"),
        ("2. 軟硬體設備資本支出", "購置 Edge AI 運算卡、本地伺服器與資料庫", "", "11-機器設備 / 18-電腦軟體"),
        ("3. 学研與法人委外", "委託高科大進行演算法優化、PMC 進行系統整合", "", "52-委外研發費 (產學合作)"),
        ("4. 消耗性器材與材料", "研發測試用零組件、雲端 API/Token 消耗費", "", "53-材料與消耗品費"),
        ("5. 員工教育訓練費", "聘請鳳凰 AI 進行主管守則研習與技術實戰內訓", "", "55-教育訓練費"),
    ]
    
    curr_row = 4
    for item in part_a_rows:
        for col_idx, val in enumerate(item, start=1):
            cell = ws1.cell(row=curr_row, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = cell_border
            if col_idx == 3:
                cell.alignment = align_center
            elif col_idx == 1:
                cell.font = bold_regular_font
                cell.alignment = align_left
            else:
                cell.alignment = align_left
            if curr_row % 2 == 1:
                cell.fill = row_alt_fill
        ws1.row_dimensions[curr_row].height = 25
        curr_row += 1
        
    # Spacer
    curr_row += 1
    
    # Part B Title
    ws1.merge_cells(f"A{curr_row}:D{curr_row}")
    ws1[f"A{curr_row}"] = "【Part B：多層次政策工具組合套利（以 NT$ 800 萬智慧檢測研發專案為例）】"
    ws1[f"A{curr_row}"].font = section_font
    ws1[f"A{curr_row}"].fill = section_fill
    ws1[f"A{curr_row}"].alignment = align_left
    ws1.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    # Part B headers
    headers_b = ["政策工具 / 戰略方案", "套利說明 / 效益金額", "", ""]
    ws1.merge_cells(f"B{curr_row}:D{curr_row}")
    for col_idx, h in enumerate(headers_b[:2], start=1):
        cell = ws1.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws1.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    part_b_rows = [
        ("戰略 1：申請「中企署 SBIR Stage 2 研發補助」", "獲得 50% 資金補貼 (NT$ 4,000,000)"),
        ("戰略 2：適用「產業創新條例第 10 條之 1」", "智慧機械與 AI 系統購置支出之 5% 抵減營所稅 (以軟硬體購置費 NT$ 2,000,000 計算，可直接抵減 NT$ 100,000 稅額)"),
        ("戰略 3：對接勞動部「充電起飛計畫」", "員工內訓講師與場地費 100% 全額補助 (NT$ 150,000)"),
        ("戰略 4：對接工研院/PMC 產學聯合加分通道", "免除內部找尋教授的行政摩擦，取得審查額外加分 5-10%"),
    ]
    
    for item in part_b_rows:
        ws1.merge_cells(f"B{curr_row}:D{curr_row}")
        cell_a = ws1.cell(row=curr_row, column=1, value=item[0])
        cell_a.font = bold_regular_font
        cell_a.alignment = align_left
        cell_a.border = cell_border
        
        cell_b = ws1.cell(row=curr_row, column=2, value=item[1])
        cell_b.font = regular_font
        cell_b.alignment = align_left_wrap
        cell_b.border = cell_border
        
        # apply border to merged cells
        for c in range(3, 5):
            ws1.cell(row=curr_row, column=c).border = cell_border
            
        ws1.row_dimensions[curr_row].height = 35
        curr_row += 1
        
    # Cost summary table
    curr_row += 1
    ws1.merge_cells(f"A{curr_row}:D{curr_row}")
    ws1[f"A{curr_row}"] = "📊 套利後實質成效精算（NT$ 800 萬專案）"
    ws1[f"A{curr_row}"].font = bold_regular_font
    ws1[f"A{curr_row}"].alignment = align_left
    ws1.row_dimensions[curr_row].height = 20
    curr_row += 1
    
    summary_headers = ["評估項目", "金額 (NT$)", "", ""]
    ws1.merge_cells(f"B{curr_row}:D{curr_row}")
    for col_idx, h in enumerate(summary_headers[:2], start=1):
        cell = ws1.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws1.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    summary_rows = [
        ("專案總預算", "8,000,000"),
        ("獲得政府無償補助款", "4,000,000"),
        ("營所稅直接抵減額", "100,000"),
        ("內訓經費無償補助", "150,000"),
        ("企業實質自籌支出", "3,750,000 (自籌比例自 100% 降至 46.8%！)"),
    ]
    
    for item in summary_rows:
        ws1.merge_cells(f"B{curr_row}:D{curr_row}")
        cell_a = ws1.cell(row=curr_row, column=1, value=item[0])
        cell_a.font = regular_font
        cell_a.alignment = align_left
        cell_a.border = cell_border
        
        cell_b = ws1.cell(row=curr_row, column=2, value=item[1])
        cell_b.alignment = align_left
        cell_b.border = cell_border
        
        if "企業實質自籌支出" in item[0]:
            cell_a.font = bold_regular_font
            cell_b.font = Font(name=font_family, size=10, bold=True, color="0D9488")
            cell_a.fill = accent_fill
            cell_b.fill = accent_fill
            for c in range(3, 5):
                ws1.cell(row=curr_row, column=c).fill = accent_fill
        else:
            cell_b.font = regular_font
            
        for c in range(3, 5):
            ws1.cell(row=curr_row, column=c).border = cell_border
            
        ws1.row_dimensions[curr_row].height = 25
        curr_row += 1
        
    # Spacer
    curr_row += 1
    
    # Part C
    ws1.merge_cells(f"A{curr_row}:D{curr_row}")
    ws1[f"A{curr_row}"] = "【Part C：產學與法人合作資源對接】"
    ws1[f"A{curr_row}"].font = section_font
    ws1[f"A{curr_row}"].fill = section_fill
    ws1[f"A{curr_row}"].alignment = align_left
    ws1.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    headers_c = ["合作單位", "對接狀態 / 說明", "", ""]
    ws1.merge_cells(f"B{curr_row}:D{curr_row}")
    for col_idx, h in enumerate(headers_c[:2], start=1):
        cell = ws1.cell(row=curr_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws1.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    part_c_rows = [
        ("高科大 AI 研發中心", "[ ] 預定合作學術機構（提供學術正統性加分）"),
        ("國立大學工學院", "[ ] 預定合作學術機構"),
        ("精密機械研究發展中心 (PMC)", "[ ] 預定法人聯合輔導窗口（中南部傳產製造業主要對接窗口）"),
        ("工業技術研究院 (工研院)", "[ ] 預定法人聯合輔導窗口"),
    ]
    
    for item in part_c_rows:
        ws1.merge_cells(f"B{curr_row}:D{curr_row}")
        cell_a = ws1.cell(row=curr_row, column=1, value=item[0])
        cell_a.font = bold_regular_font
        cell_a.alignment = align_left
        cell_a.border = cell_border
        
        cell_b = ws1.cell(row=curr_row, column=2, value=item[1])
        cell_b.font = regular_font
        cell_b.alignment = align_left
        cell_b.border = cell_border
        
        for c in range(3, 5):
            ws1.cell(row=curr_row, column=c).border = cell_border
            
        if curr_row % 2 == 1:
            cell_a.fill = row_alt_fill
            cell_b.fill = row_alt_fill
            for c in range(3, 5):
                ws1.cell(row=curr_row, column=c).fill = row_alt_fill
                
        ws1.row_dimensions[curr_row].height = 25
        curr_row += 1
        
    # ---------------------------------------------------------
    # Sheet 2: 📊 補助匹配評估與風險控制矩陣
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="📊 補助匹配與風險矩陣")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "補助匹配評估與風險控制矩陣"
    ws2["A1"].font = title_font
    ws2["A1"].fill = title_fill
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 40
    
    headers_2 = ["評估維度 / 指標細節", "計畫 A (如 SBIR) 分數", "計畫 B (如 智慧機械) 分數", "實務評估指引與風險控制紅線"]
    for col_idx, h in enumerate(headers_2, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws2.row_dimensions[2].height = 25
    
    matrix_rows = [
        ("1. 資金自籌與流動性 (Liquidity & Cash Flow)\n- 公司是否有能力墊付前期的自籌款？\n- 能否承受補助款「分期撥付」的時差壓力？", "", "", "⚠️ 紅線：公司帳上可用現金儲備必須大於「自籌款總額 * 1.5 倍」，否則極易因週轉不靈出局。"),
        ("2. 研發團隊穩定性與工時合規 (Team Continuity)\n- 核心演算法工程師流動風險是否 < 25%？\n- 能否確保研發紀錄 100% 留存 Git & Jira log？", "", "", "⚠️ 紅線：研發人員近半年離職率超過 25% 者不建議申請，因變更人員會面臨極嚴苛的審查與核銷。"),
        ("3. 關鍵工作里程碑合理性 (Milestone Feasibility)\n- 開發時程 (12-24個月) 是否符合產品路線圖？\n- 是否有足夠時間進行 AI 模型迭代與 Retraining？", "", "", "⚠️ 指引：執行期建議設定在 12 至 18 個月，預留至少 2 個月作為系統整合與測試報告緩衝期。"),
        ("4. 技術量化 KPI 結案難易度 (KPI Audit Compliance)\n- KPI 指標是否純屬技術性 (如 Precision, Latency)？\n- 是否避開受外部市場波動影響的業績型 KPI？", "", "", "⚠️ 紅線：嚴禁編列「增加營業額 XX 萬」等市場型 KPI。必須採用 80% 安全防護欄公式進行編列。"),
        ("5. 產學與外協合作摩擦度 (Collaboration Friction)\n- 學研合約的 IP 智財權歸屬是否已達成共識？\n- 委外開發商是否具備 AI 能量登錄白名單？", "", "", "⚠️ 指引：評估大學教授合約談判時程是否會拖累計畫。若有困難可優先考慮鳳凰 AI 的法人資源。"),
        ("6. 資安與法規合規成本 (Security & Legal Overhead)\n- 開發流程是否需符合個資法第 8 條告知規範？\n- 系統是否需留存詳細的審計日誌 (Audit Trail)？", "", "", "⚠️ 指引：是否編列 ISO/IEC 42001 與個資防護的合規預算？此項為 2026 年後審查委員的加分關鍵。")
    ]
    
    curr_row = 3
    for item in matrix_rows:
        ws2.cell(row=curr_row, column=1, value=item[0]).alignment = align_left_wrap
        ws2.cell(row=curr_row, column=2, value=item[1]).alignment = align_center
        ws2.cell(row=curr_row, column=3, value=item[2]).alignment = align_center
        ws2.cell(row=curr_row, column=4, value=item[3]).alignment = align_left_wrap
        
        ws2.cell(row=curr_row, column=1).font = bold_regular_font
        ws2.cell(row=curr_row, column=2).font = regular_font
        ws2.cell(row=curr_row, column=3).font = regular_font
        ws2.cell(row=curr_row, column=4).font = regular_font
        
        # Check for red warnings in red font
        if "紅線" in item[3]:
            ws2.cell(row=curr_row, column=4).font = alert_font
            ws2.cell(row=curr_row, column=4).fill = alert_fill
            
        for c in range(1, 5):
            ws2.cell(row=curr_row, column=c).border = cell_border
            if curr_row % 2 == 1 and not "紅線" in item[3]:
                ws2.cell(row=curr_row, column=c).fill = row_alt_fill
                
        ws2.row_dimensions[curr_row].height = 65
        curr_row += 1
        
    # Total row
    ws2.cell(row=curr_row, column=1, value="📊 總分累計 (滿分 30 分)").font = bold_regular_font
    ws2.cell(row=curr_row, column=1).alignment = align_left
    ws2.cell(row=curr_row, column=2, value="[ 手動填寫 ]").alignment = align_center
    ws2.cell(row=curr_row, column=3, value="[ 手動填寫 ]").alignment = align_center
    ws2.cell(row=curr_row, column=4, value="💡 決策矩陣評分指引見下方").font = regular_font
    ws2.cell(row=curr_row, column=4).alignment = align_left
    
    ws2.cell(row=curr_row, column=1).fill = section_fill
    ws2.cell(row=curr_row, column=2).fill = section_fill
    ws2.cell(row=curr_row, column=3).fill = section_fill
    ws2.cell(row=curr_row, column=4).fill = section_fill
    
    for c in range(1, 5):
        ws2.cell(row=curr_row, column=c).border = cell_border
        ws2.cell(row=curr_row, column=c).font = bold_regular_font
        
    ws2.row_dimensions[curr_row].height = 30
    curr_row += 2
    
    # Guidelines section
    ws2.merge_cells(f"A{curr_row}:D{curr_row}")
    ws2[f"A{curr_row}"] = "💡 決策矩陣評分指引與行動計畫"
    ws2[f"A{curr_row}"].font = section_font
    ws2[f"A{curr_row}"].fill = section_fill
    ws2[f"A{curr_row}"].alignment = align_left
    ws2.row_dimensions[curr_row].height = 25
    curr_row += 1
    
    guideline_items = [
        ("● 總分 >= 24 分", "【極佳匹配】項目可行性極高，專案團隊可即刻著手撰寫計畫書並進行申報！"),
        ("● 18 - 23 分", "【黃色預警】部分指標存在潛在風險（如現金流或工時）。建議修正 KPI 指標、調整預算編列或縮短執行期，再送件。"),
        ("● < 18 分", "【紅色否決】風險過高（如團隊不穩或自籌款嚴重不足）。極易結案失敗或被追回補助款，不建議在現有架構下申請。")
    ]
    
    for title, desc in guideline_items:
        ws2.merge_cells(f"B{curr_row}:D{curr_row}")
        cell_a = ws2.cell(row=curr_row, column=1, value=title)
        cell_a.font = bold_regular_font
        cell_a.alignment = align_left
        cell_a.border = cell_border
        
        cell_b = ws2.cell(row=curr_row, column=2, value=desc)
        cell_b.font = regular_font
        cell_b.alignment = align_left_wrap
        cell_b.border = cell_border
        
        for c in range(3, 5):
            ws2.cell(row=curr_row, column=c).border = cell_border
            
        if "紅色" in desc:
            cell_a.fill = alert_fill
            cell_b.fill = alert_fill
            for c in range(3, 5):
                ws2.cell(row=curr_row, column=c).fill = alert_fill
            cell_a.font = alert_font
            cell_b.font = alert_font
        elif "黃色" in desc:
            warn_fill = PatternFill(fill_type="solid", start_color="FFFBEB", end_color="FFFBEB") # Light amber tint
            warn_font = Font(name=font_family, size=10, color="92400E")
            cell_a.fill = warn_fill
            cell_b.fill = warn_fill
            for c in range(3, 5):
                ws2.cell(row=curr_row, column=c).fill = warn_fill
            cell_a.font = warn_font
            cell_b.font = warn_font
        else:
            cell_a.fill = accent_fill
            cell_b.fill = accent_fill
            for c in range(3, 5):
                ws2.cell(row=curr_row, column=c).fill = accent_fill
            cell_a.font = accent_font
            cell_b.font = accent_font
            
        ws2.row_dimensions[curr_row].height = 30
        curr_row += 1
        
    # ---------------------------------------------------------
    # Sheet 3: ✍️ 計畫書 7 大核心骨架實務手冊
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="✍️ 計畫書 7 大核心骨架")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "計畫書 7 大核心骨架與撰寫實務指南"
    ws3["A1"].font = title_font
    ws3["A1"].fill = title_fill
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 40
    
    # Tips Box
    ws3.merge_cells("A2:D4")
    tips_text = (
        "💡 孟淑慧指導講師專欄：審查委員絕不公開的 3 大評審潛規則\n"
        "1. 潛規則 1 (8分鐘掃讀)：委員每天審數十份計畫書，每份僅停留 10 分鐘。必須首頁結論先行，並繪製一目了然的系統架構圖。\n"
        "2. 潛規則 2 (產學虛與實)：學校要論文，企業要落地。建議將陳策略長或孟顧問列為專家委員，兼顧加分與研發進度。\n"
        "3. 潛規則 3 (KPI 80% 安全防護)：效益指標（如良率提升）承諾實際預估值的 80% 即可，確保結案順利甚至超標，防範被追回款項。"
    )
    ws3["A2"] = tips_text
    ws3["A2"].font = Font(name=font_family, size=9.5, color="1E3A8A")
    ws3["A2"].fill = PatternFill(fill_type="solid", start_color="EFF6FF", end_color="EFF6FF") # Light blue tint
    ws3["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in range(2, 5):
        for c in range(1, 5):
            ws3.cell(row=r, column=c).border = cell_border
            if r > 2 or c > 1:
                ws3.cell(row=r, column=c).fill = PatternFill(fill_type="solid", start_color="EFF6FF", end_color="EFF6FF")
    ws3.row_dimensions[2].height = 25
    ws3.row_dimensions[3].height = 25
    ws3.row_dimensions[4].height = 30
    
    # Headers
    headers_3 = ["骨架編號與名稱", "委員關注點", "撰寫公式", "傳統扣件廠實務範例"]
    for col_idx, h in enumerate(headers_3, start=1):
        cell = ws3.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws3.row_dimensions[5].height = 25
    
    skeleton_rows = [
        (
            "骨架 1：企業背景與轉型痛點\n(Background & Bottlenecks)",
            "為什麼是這家企業需要補助？痛點是否具備產業代表性？有沒有具體損失數據？",
            "現有營運模式 + 人工瓶頸定量數據 + 帶來的財務/客戶損失 = 政策補助介入的必要性。",
            "本公司為擁有 35 年歷史的精密機車螺絲扣件製造商。現行品質檢驗 100% 依賴人工目視，然而因扣件微小瑕疵（如螺牙損傷、裂痕）特徵複雜，人工檢驗平均漏檢率高達 2.8%。此瓶頸導致 2025 年遭歐洲 Tier-1 車廠客戶罰款與退貨損失達新台幣 450 萬元，並嚴重威脅台廠在國際供應鏈中的信譽。因高階 AI 研發與邊緣設備購置成本極高，亟需政府 SBIR 計畫扶持，以分擔早期技術轉型風險。"
        ),
        (
            "骨架 2：研發目標與關鍵技術創新點\n(R&D Goals & Core Innovations)",
            "這是否只是單純的硬體採購或 SaaS 軟體租用？核心研發難點在哪？創新性是否顯著？",
            "解構通用SaaS採購 + 自主研發/微調的核心演算法 (如 LoRA/RAG/Edge AI) + 解決的特定技術難題。",
            "本計畫非單純採購現成光學檢測機台，而是研發一套「邊緣運算多模態 AI 扣件瑕疵分類系統」。我們將自主開發針對金屬反光雜訊的預處理濾波演算法，並採用 YOLO-v11 進行毫秒級瑕疵定位，結合本地部署之 Llama-3-8B-Instruct 進行瑕疵原因多模態推理與生成。針對扣件瑕疵特徵進行「LoRA (Low-Rank Adaptation) 精準微調」，克服傳統 CV 難以識別微小隱性裂痕的技術難題，實現 100% 台灣自主產權。"
        ),
        (
            "骨架 3：系統架構設計與 AI 演算法流程\n(System Architecture & Pipeline)",
            "技術架構是否清晰合理？是否有考量到 MLOps 生命週期與資料漂移？資料流與安全如何確保？",
            "邊緣採購與資料流 ➔ 向量資料庫/特徵提取 ➔ 模型微調與推理 ➔ 可解釋性審計軌跡 ➔ MLOps 重新訓練閉環。",
            "本系統架裝包含三大層級：(1) 數據感知層：產線工業相機毫秒級抓拍扣件影像，並透過本地網閘進行遮罩去識別化；(2) 演算法推理層：將影像特徵送入地端伺服器之 pgvector 特徵庫，執行 YOLO 瑕疵定位與 LLM 原因推理；(3) MLOps 治理層：系統配置資料漂移監控模組。一旦產線更換扣件型號（Data Drift），系統將自動觸發 Active Learning 機制篩選高價值樣本，啟動 LoRA 自動重新訓練（Retraining Pipeline），並寫入符合 ISO/IEC 42001 規範之技術日誌。"
        ),
        (
            "骨架 4：關鍵工作項目與開發里程碑\n(Key Milestones & Gantt)",
            "開發時程安排是否合理？每個里程碑的「交付成果」是否具備可稽核性 (Auditability)？",
            "將計畫拆解為四個階段（D1-D4），明確列出時間點與查核實物（如設計書、代碼庫、測試報告）。",
            "第一階段 (D1, 第 1-3 個月) 系統分析與設計：交付《AI 瑕疵分類系統架構規格書》與《資料集標記規範手冊》。\n第二階段 (D2, 第 4-8 個月) 演算法研發與離線訓練：完成 YOLO-v11 與本地 LLM 微調，交付《演算法模型評估報告（Precision 達 92% 驗證）》與原始代碼庫。\n第三階段 (D3, 第 9-12 個月) 系統整合與 Alpha 測試：將 AI 演算法與產線硬體/PLC 控制器完成串接，交付《產線整合測試報告（時延 < 80ms 驗證）》。\n第四階段 (D4, 第 13-16 個月) 現場 Beta 測試與結案準備：進行產線 100,000次實際檢測，交付《第三方公正單位驗證報告（如 PMC 瑕疵檢測率認證）》與《結案自評報告》。"
        ),
        (
            "骨架 5：人力資源分配與研發工時安排\n(Personnel & Compliance)",
            "研發人員配置是否足夠？有沒有幽靈人口？是否符合台灣勞基法規範？",
            "專案負責人 (高管) + 研發主管 (PM) + 核心工程師 + 外部專家/學研顧問 + 100%符合單月加班 < 46小時與實質 Git log 對位。",
            "配置專案負責人（總經理特助，投入工時 30%）、研發經理（資深軟體主管，投入工時 60%）、AI 工程師（2名，投入工時 80%）、資料標記員（1名，投入工時 50%）。聘請高科大陳文家策略長與鳳凰 AI 孟淑慧首席顧問擔任技術顧問（投入工時 15%）。所有內部研發人員皆需強制使用 Git 進行代碼提交與簽到退，工時表與薪資扣繳憑單 100% 吻合，絕無虛報，且嚴格遵守勞基法單日總工時不超 12 小時、單月加班不超 46 小時之底線。"
        ),
        (
            "骨架 6：經費預算配比與會計科目規劃\n(Budget & Accounting)",
            "會計科目編列是否合規？有沒有將資本支出誤列為經常費用？是否規劃了稅額抵減？",
            "人事費 50-60% + 學研委外費 < 30% + 消耗性器材費 10-20% + 設備折舊與產創 10-1 租稅抵減規畫。",
            "專案總預算編列新台幣 800 萬元。其中：(1) 研發人事費：編列 NT$ 440 萬（55%，符合人事費不超 60% 限制）；(2) 消耗性器材與材料費：編列 NT$ 80 萬（10%，用於雲端 GPU 訓練 Token 費與測試模具耗損）；(3) 學研與委外開發費：委託高科大產學合作編列 NT$ 160 萬（20%）；(4) 研發設備使用折舊費：購置地端伺服器與 Edge AI 卡編列 NT$ 120 萬（15%）。此購置設備同時向財政部申請「產業創新條例第 10 條之 1」智慧機械與 AI 系統抵減，預計可享 5% 營所稅抵減（NT$ 6 萬元），實現多重政策套利。"
        ),
        (
            "骨架 7：預期效益與量化/質化 KPI 指標\n(KPIs & Safety Margins)",
            "指標是否具備可測量性 (Measurability)？結案時如何查核？是否吹牛過度導致結案翻船？",
            "產出型 KPI (必定能完成) + 效益型 KPI (套用 80% 安全防護欄公式) + 質化合規指標 (ISO/IEC 42001 / 個資政策)。",
            "1. 產出型量化 KPI：開發邊緣運算 AI 扣件瑕疵檢測系統 1 套；完成 30,000 張螺絲螺帽瑕疵標記資料集 1 份；申請專利 1 件。\n2. 效益型量化 KPI (80% 安全防護欄)：\n- 預期產線實際提升效率 95% ➔ 計畫書僅承諾 76%（95% * 80%）。\n- 預期瑕疵檢測時間由 5.0 秒降至 0.8 秒 ➔ 計畫書僅承諾降至 1.5 秒。\n- 預期漏檢率降至 0.2% ➔ 計畫書僅承諾降至 < 0.5%。\n3. 質化效益 KPI：依據 ISO/IEC 42001 框架建立《扣件產線 AI 演算法透明度與安全稽核作業規範書》；建立符合台灣個資法第 8 條之顧客專屬宣告政策 1 份。"
        )
    ]
    
    curr_row = 6
    for item in skeleton_rows:
        ws3.cell(row=curr_row, column=1, value=item[0]).alignment = align_left_wrap
        ws3.cell(row=curr_row, column=2, value=item[1]).alignment = align_left_wrap
        ws3.cell(row=curr_row, column=3, value=item[2]).alignment = align_left_wrap
        ws3.cell(row=curr_row, column=4, value=item[3]).alignment = align_left_wrap
        
        ws3.cell(row=curr_row, column=1).font = bold_regular_font
        ws3.cell(row=curr_row, column=2).font = regular_font
        ws3.cell(row=curr_row, column=3).font = regular_font
        ws3.cell(row=curr_row, column=4).font = regular_font
        
        for c in range(1, 5):
            ws3.cell(row=curr_row, column=c).border = cell_border
            if curr_row % 2 == 1:
                ws3.cell(row=curr_row, column=c).fill = row_alt_fill
                
        ws3.row_dimensions[curr_row].height = 110
        curr_row += 1
        
    # ---------------------------------------------------------
    # Sheet 4: 📅 月度執行與風險追蹤表
    # ---------------------------------------------------------
    ws4 = wb.create_sheet(title="📅 月度執行與風險追蹤")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "補助案月度執行與合規風險追蹤表"
    ws4["A1"].font = title_font
    ws4["A1"].fill = title_fill
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 40
    
    headers_4 = ["自核結果 (Y/N/NA)", "稽核模組與檢核要點", "實際執行情況與說明"]
    for col_idx, h in enumerate(headers_4, start=1):
        cell = ws4.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws4.row_dimensions[2].height = 25
    
    checklist_rows = [
        # Module 1
        ("SECTION", "【模組 1：技術里程碑與 Gantt Chart 進度查核】", ""),
        ("[  ]", "查核本月預定工作項目與計畫進度：是否有落後？\n- 技術產出物是否已存入內部 Git 版本控制系統，並留存 Commits 軌跡？", ""),
        ("[  ]", "若涉及委外學研機構，本月是否已取得階段性研究報告、會議紀錄或簽收紀錄？", ""),
        ("[  ]", "計畫整體進度落後是否大於 10%？\n- 若是，是否已啟動計畫變更程序或由外部專家顧問團隊介入排除？", ""),
        
        # Module 2
        ("SECTION", "【模組 2：台灣勞基法與研發工時防護欄 (🚨 審計稽核最大紅線)】", ""),
        ("[  ]", "查核本月所有計畫參與人員之「刷卡出勤紀錄」與「補助案工時表」：兩者時間是否 100% 吻合？", ""),
        ("[  ]", "經確認，無任何研發人員單日總工時超過 12 小時，且單月總加班工時皆低於 46 小時法定上限？", ""),
        ("[  ]", "本月請假者（如特休、病假、事假）是否已在工時表中扣除假別時數？（嚴禁請假當天仍申報研發工時）", ""),
        ("[  ]", "所有研發日誌 (R&D Log) 的工作描述，是否皆為「具體技術開發與實驗內容」，而非行政行政雜務？", ""),
        
        # Module 3
        ("SECTION", "【模組 3：AI/MLOps 模型健康與資料合規追蹤】", ""),
        ("[  ]", "本月產線資料或影像特徵值是否發生偏移 (Data Drift)？\n- 若是，是否已觸發重新訓練以維持模型精度？", ""),
        ("[  ]", "地端 AI 系統之安全防護日誌中，是否有 Prompt Injection 或惡意攻擊嘗試記錄？\n- 是否已依 ISO/IEC 42001 安全規範進行 IP 封鎖或熔斷？", ""),
        ("[  ]", "AI 推理結果與人工覆核紀錄之偏差率是否維持在安全邊界內（如目標值 < 1.5%）？", "")
    ]
    
    curr_row = 3
    for item in checklist_rows:
        if item[0] == "SECTION":
            ws4.merge_cells(f"A{curr_row}:C{curr_row}")
            cell = ws4.cell(row=curr_row, column=1, value=item[1])
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = align_left
            for c in range(1, 4):
                ws4.cell(row=curr_row, column=c).border = cell_border
            ws4.row_dimensions[curr_row].height = 25
        else:
            cell_a = ws4.cell(row=curr_row, column=1, value=item[0])
            cell_a.font = bold_regular_font
            cell_a.alignment = align_center
            cell_a.border = cell_border
            
            cell_b = ws4.cell(row=curr_row, column=2, value=item[1])
            cell_b.font = regular_font
            cell_b.alignment = align_left_wrap
            cell_b.border = cell_border
            
            cell_c = ws4.cell(row=curr_row, column=3, value=item[2])
            cell_c.font = regular_font
            cell_c.alignment = align_left
            cell_c.border = cell_border
            
            if "🚨" in item[1] or "紅線" in item[1]:
                cell_b.font = Font(name=font_family, size=10, bold=True, color="991B1B")
                cell_b.fill = alert_fill
                cell_a.fill = alert_fill
                cell_c.fill = alert_fill
            elif curr_row % 2 == 1:
                cell_a.fill = row_alt_fill
                cell_b.fill = row_alt_fill
                cell_c.fill = row_alt_fill
                
            ws4.row_dimensions[curr_row].height = 40
            
        curr_row += 1
        
    # ---------------------------------------------------------
    # Sheet 5: 📋 結案核銷自評檢核表
    # ---------------------------------------------------------
    ws5 = wb.create_sheet(title="📋 結案核銷 20 點自評檢核")
    ws5.views.sheetView[0].showGridLines = True
    
    ws5.merge_cells("A1:D1")
    ws5["A1"] = "結案實地審查與核銷 20 點自評檢核表 (會計師進場前 30 天必做)"
    ws5["A1"].font = title_font
    ws5["A1"].fill = title_fill
    ws5["A1"].alignment = align_center
    ws5.row_dimensions[1].height = 40
    
    headers_5 = ["自核結果 (Y/N/NA)", "項目編號", "稽核範疇與檢核點", "對應佐證憑證 / 核銷說明"]
    for col_idx, h in enumerate(headers_5, start=1):
        cell = ws5.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
    ws5.row_dimensions[2].height = 25
    
    checklist_20 = [
        # Group 1
        ("SECTION", "一、財務、發票與會計科目合規 (Financial & Invoices)", "", ""),
        ("[  ]", "1", "計畫內所有採購發票之「開立日期」，皆落在計畫執行期 (自 Day 1 至最後一日) 之內。", "發票正本及影本"),
        ("[  ]", "2", "所有發票之品名、規格，與計畫書中編列的設備明細完全一致 (若不同，需已核准變更)。", "統一發票、財產清冊、驗收單"),
        ("[  ]", "3", "預算科目間之流用 (如人事費流用至材料費) 比例符合規定 (一般無事前申請不得超過 20%)。", "預算流用變更申請書"),
        ("[  ]", "4", "所有採購皆備妥三家廠商之「比價單/估價單」與開標紀錄 (符合政府採購法或計畫核銷規範)。", "估價單三張、比價紀錄表"),
        ("[  ]", "5", "企業自籌款實際支付憑證 (公司存摺影本、銀行匯款單) 100% 備妥，且無異常退款。", "銀行存摺影本、轉帳單"),
        
        # Group 2
        ("SECTION", "二、人力資源、薪資與工時憑證 (HR & Attendance)", "", ""),
        ("[  ]", "6", "所有參與計畫員工之「勞保投保明細」及「健保投保紀錄」皆在案，且投保薪資與申報薪資吻合。", "勞保投保明細、健保投保紀錄"),
        ("[  ]", "7", "計畫執行期間之「薪資轉帳存摺影本」及「薪資單/薪資明細表」100% 備妥。", "薪資發放憑證、劃撥單"),
        ("[  ]", "8", "年度申報之「扣繳憑單 (50-薪資申報)」中，該研發人員申報金額與計畫核銷金額完全對位。", "年度扣繳憑單影本"),
        ("[  ]", "9", "研發工時表之親筆簽核 (或合規電子簽章) 完整，無漏簽或代理代簽現象。", "研發簽到退工時表"),
        ("[  ]", "10", "經會計稽核，無 any 非計畫編列人員之工時被誤報入本專案。", "會計核銷總表"),
        
        # Group 3
        ("SECTION", "三、設備、財產登記與資安防護 (Equipment & Security)", "", ""),
        ("[  ]", "11", "專案經費購置之伺服器、Edge 設備，皆已張貼「經濟部 [計畫名稱] 補助購置」之金屬財產標籤。", "設備照片、財產登記卡"),
        ("[  ]", "12", "購置之軟硬體設備皆已登錄於公司財產目錄，且折舊年限及殘值計算符合所得稅法規定。", "財產目錄、折舊清冊"),
        ("[  ]", "13", "設備安裝地點與計畫書所列「研發場地」完全一致，無私自移動至非授權場所或供私人使用之情事。", "設備放置地點照片與配置圖"),
        ("[  ]", "14", "系統已配置符合 ISO/IEC 42001 規範之存取控制權限，且研發用機密資料皆有去識別化遮罩紀錄。", "權限管理設定、資料遮罩規格書"),
        ("[  ]", "15", "留存至少 6 個月之系統備份紀錄與資安稽核日誌 (Security Logs)，供審查委員隨機抽查。", "備份日誌、系統稽核日誌"),
        
        # Group 4
        ("SECTION", "四、技術 KPI 指標、成果展示與佐證 (Technical KPIs & Demo)", "", ""),
        ("[  ]", "16", "量化 KPI 所承諾之數據 (如瑕疵檢測率 > 95%)，已取得第三方公正單位 (如 PMC/資策會) 之測試報告。", "公正單位測試/驗證報告"),
        ("[  ]", "17", "系統開發成果之「原始代碼倉庫 (Git Repository)」版本演進紀錄完整，無結案前一次性上傳之嫌疑。", "Git Commit Log、代碼清冊"),
        ("[  ]", "18", "計畫承諾之「系統規格書」、「測試報告」、「使用者操作手冊」等技術文檔皆已裝訂成冊。", "規格書、說明手冊紙本"),
        ("[  ]", "19", "現場展示 (Live Demo) 產線動態流程已演練完畢，且已預先錄製「備用展示影片」，防範現場斷網突發。", "Demo備用展示影片檔"),
        ("[  ]", "20", "依個資法第 8 條規定跳出之「個資告知同意書」線上實裝畫面截圖與資料遮罩存檔已準備就緒。", "個資聲明截圖、資料流向證明")
    ]
    
    curr_row = 3
    for item in checklist_20:
        if item[0] == "SECTION":
            ws5.merge_cells(f"A{curr_row}:D{curr_row}")
            cell = ws5.cell(row=curr_row, column=1, value=item[1])
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = align_left
            for c in range(1, 5):
                ws5.cell(row=curr_row, column=c).border = cell_border
            ws5.row_dimensions[curr_row].height = 25
        else:
            cell_a = ws5.cell(row=curr_row, column=1, value=item[0])
            cell_a.font = bold_regular_font
            cell_a.alignment = align_center
            cell_a.border = cell_border
            
            cell_b = ws5.cell(row=curr_row, column=2, value=item[1])
            cell_b.font = bold_regular_font
            cell_b.alignment = align_center
            cell_b.border = cell_border
            
            cell_c = ws5.cell(row=curr_row, column=3, value=item[2])
            cell_c.font = regular_font
            cell_c.alignment = align_left_wrap
            cell_c.border = cell_border
            
            cell_d = ws5.cell(row=curr_row, column=4, value=item[3])
            cell_d.font = regular_font
            cell_d.alignment = align_left_wrap
            cell_d.border = cell_border
            
            if curr_row % 2 == 1:
                cell_a.fill = row_alt_fill
                cell_b.fill = row_alt_fill
                cell_c.fill = row_alt_fill
                cell_d.fill = row_alt_fill
                
            ws5.row_dimensions[curr_row].height = 35
            
        curr_row += 1

    # ---------------------------------------------------------
    # Auto-adjust column widths
    # ---------------------------------------------------------
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Avoid merged cells scaling the column width excessively
                if cell.coordinate in ws.merged_cells:
                    # check if it's the top-left cell of the merge
                    is_top_left = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            is_top_left = True
                            break
                    if not is_top_left:
                        continue
                if cell.value:
                    val_str = str(cell.value)
                    # Simple length estimate: count Chinese characters as 2
                    cell_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                    if cell_len > max_len:
                        max_len = cell_len
            # Apply width with safety margin
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)
            
    # Save file
    target_dir = "curriculum/unit_8_grants"
    os.makedirs(target_dir, exist_ok=True)
    target_path = os.path.join(target_dir, "phoenix_ai_grants_handbook.xlsx")
    wb.save(target_path)
    print(f"Excel 檔案已成功生成並存檔至: {target_path}")

if __name__ == "__main__":
    create_handbook()
