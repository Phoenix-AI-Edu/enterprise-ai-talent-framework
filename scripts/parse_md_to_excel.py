# -*- coding: utf-8 -*-
"""
parse_md_to_excel.py — 將 Markdown 專案管理表與 A3 戰略畫布轉換為 Excel (.xlsx) 檔案
========================================================================
[Confidential - Phoenix AI Internal Asset]

本程式提供將 Markdown 格式的 4 大 Notion 資料表自動解析並轉存為實體 Excel 的功能。
支援：
- 自動偵測 H2 標題做為工作表 (Sheet) 名稱，並將其對齊為乾淨的英文命名。
- 解析標準 Markdown 表格，自動過濾對齊線 (e.g. | :--- |)。
- 偵測並解析 H2/H3/代碼塊 中的 A3 戰略畫布 (A3 Strategy Canvas)，將其轉換為第 5 個工作表。
- 使用 openpyxl 進行專業樣式套用 (自適應欄寬最大 65、表頭深藍底白字、框線設定)。
"""

import os
import sys
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_markdown_cell(text):
    """清理單元格中的 Markdown 語法，例如去除連結或加粗，保留乾淨文字"""
    clean_text = text.strip()
    # 移除連結語法 [顯示文字](連結網址) -> 顯示文字
    clean_text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', clean_text)
    # 移除加粗 **文字** -> 文字
    clean_text = re.sub(r'\*\*([^*]+)\*\*', r'\1', clean_text)
    return clean_text

def map_sheet_title(raw_title):
    """將 Markdown 標題對應到乾淨的 Excel 工作表名稱"""
    raw_title_clean = raw_title.lower()
    if "scenario" in raw_title_clean or "場景" in raw_title_clean:
        return "Scenario Inventory"
    elif "architecture" in raw_title_clean or "架構" in raw_title_clean or "資料基礎" in raw_title_clean:
        return "Data Architecture"
    elif "adoption" in raw_title_clean or "試點" in raw_title_clean:
        return "Pilot Adoption"
    elif "governance" in raw_title_clean or "營運" in raw_title_clean or "治理" in raw_title_clean:
        return "Operations Governance"
    elif "canvas" in raw_title_clean or "戰略畫布" in raw_title_clean or "總表" in raw_title_clean:
        return "A3 Strategy Canvas"
    
    # 去除表情符號與資料表前綴以簡化 Sheet 名稱
    clean_title = re.sub(r'[^\w\s\u4e00-\u9fa5]', '', raw_title)
    clean_title = clean_title.replace("資料表一", "").replace("資料表二", "").replace("資料表三", "").replace("資料表四", "")
    clean_title = clean_title.replace("企業 AI ", "").replace("資料庫", "").strip()
    # 限制長度在 31 字元內 (Excel Sheet 名稱限制)
    return clean_title[:30] if clean_title else "Sheet"

def parse_a3_canvas_from_text(text_block):
    """從 6 盒戰略畫布的文字框格中，抽取各區塊與其內容"""
    lines = text_block.strip().split("\n")
    # 過濾出含有框線字元的列
    canvas_lines = []
    for line in lines:
        if any(c in line for c in "┌┐└┘├┤┬┴┼─│"):
            canvas_lines.append(line)
            
    if not canvas_lines:
        return []

    # 依分隔線進行分群
    groups = []
    current_group = []
    for line in canvas_lines:
        is_sep = False
        if line.strip().startswith("├") or line.strip().startswith("┌") or line.strip().startswith("└"):
            is_sep = True
        
        if is_sep:
            if current_group:
                groups.append(current_group)
                current_group = []
        else:
            current_group.append(line)
    if current_group:
        groups.append(current_group)

    # 定義 7 個區塊：1-6 盒 + 終極目標
    boxes = {i: {"title": "", "content": []} for i in range(1, 8)}
    
    def clean_cell(text):
        return text.strip()

    # 解析第一群 (1, 2, 3 盒)
    if len(groups) > 1:
        g1 = groups[1]
        first_line = g1[0]
        parts = [clean_cell(p) for p in first_line.split("│")[1:-1]]
        for idx, title in enumerate(parts):
            if idx < 3:
                boxes[idx+1]["title"] = title
                
        for line in g1[1:]:
            parts = [clean_cell(p) for p in line.split("│")[1:-1]]
            for idx, part in enumerate(parts):
                if idx < len(parts) and idx < 3:
                    part_clean = part.strip()
                    if part_clean and not part_clean.startswith("(對應"):
                        boxes[idx+1]["content"].append(part_clean)

    # 解析第二群 (4, 5, 6 盒)
    if len(groups) > 2:
        g2 = groups[2]
        first_line = g2[0]
        parts = [clean_cell(p) for p in first_line.split("│")[1:-1]]
        for idx, title in enumerate(parts):
            if idx < 3:
                boxes[idx+4]["title"] = title
                
        for line in g2[1:]:
            parts = [clean_cell(p) for p in line.split("│")[1:-1]]
            for idx, part in enumerate(parts):
                if idx < len(parts) and idx < 3:
                    part_clean = part.strip()
                    if part_clean and not part_clean.startswith("(對應"):
                        boxes[idx+4]["content"].append(part_clean)

    # 解析第三群 (終極戰略目標)
    if len(groups) > 3:
        g3 = groups[3]
        for line in g3:
            parts = [clean_cell(p) for p in line.split("│")[1:-1]]
            if parts:
                text = parts[0].strip()
                if "終極戰略目標" in text:
                    if "：" in text:
                        title, content = text.split("：", 1)
                        boxes[7]["title"] = clean_cell(title)
                        boxes[7]["content"].append(clean_cell(content))
                    else:
                        boxes[7]["title"] = "終極戰略目標"
                        boxes[7]["content"].append(text)

    # 組合區塊內容為單一文字段落，提升 Excel 可讀性
    rows = []
    for i in range(1, 8):
        box = boxes[i]
        title = box["title"]
        title = re.sub(r'[^\w\s\u4e00-\u9fa5\.\-\:\(\)]', '', title).strip()
        if not title and i == 7:
            title = "終極戰略目標"
        
        content_items = []
        current_header = ""
        current_bullets = []
        
        for item in box["content"]:
            is_header = False
            is_bullet = False
            if item.startswith("*"):
                is_header = True
            elif item.startswith("-"):
                is_bullet = True
            else:
                if item.endswith(":") or item.endswith("："):
                    is_header = True
                else:
                    is_bullet = True
            
            clean_item = re.sub(r'^[\*\-\s]+', '', item).strip()
            if not clean_item:
                continue
                
            if is_header:
                if current_bullets:
                    if current_header:
                        content_items.append(f"{current_header}{'、'.join(current_bullets)}")
                    else:
                        content_items.append("、".join(current_bullets))
                    current_bullets = []
                current_header = clean_item
            elif is_bullet:
                current_bullets.append(clean_item)
            else:
                if current_bullets:
                    if current_header:
                        content_items.append(f"{current_header}{'、'.join(current_bullets)}")
                    else:
                        content_items.append("、".join(current_bullets))
                    current_bullets = []
                    current_header = ""
                content_items.append(clean_item)
                
        if current_bullets:
            if current_header:
                content_items.append(f"{current_header}{'、'.join(current_bullets)}")
            else:
                content_items.append("、".join(current_bullets))
                
        content_str = "；".join(content_items)
        rows.append([title, content_str])
        
    return rows

def parse_a3_from_sections(filepath):
    """Fallback 方案：從 Markdown 文件的 H2/H3 結構章節中解析 A3 Canvas 內容"""
    if not os.path.exists(filepath):
        return []
        
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()
        
    sections = []
    current_section_title = ""
    current_section_content = []
    
    for line in lines:
        line_str = line.strip()
        # 偵測大綱標題如 "## 一、 ..."
        heading_match = re.match(r'^##\s+(?:[^\w\s]*\s*)?([一二三四五六七八九十]+[、\.\s].*)$', line_str)
        if heading_match:
            if current_section_title:
                sections.append((current_section_title, current_section_content))
            current_section_title = heading_match.group(1).strip()
            current_section_content = []
        else:
            if current_section_title:
                if line_str and not line_str.startswith("```"):
                    clean_line = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', line_str)
                    clean_line = re.sub(r'\*\*([^*]+)\*\*', r'\1', clean_line)
                    clean_line = re.sub(r'^[\*\-\#\s\>]+', '', clean_line).strip()
                    if clean_line:
                        current_section_content.append(clean_line)
                        
    if current_section_title:
        sections.append((current_section_title, current_section_content))
        
    rows = []
    for title, content_list in sections:
        title_clean = re.sub(r'\s*\(.*?\)', '', title).strip()
        title_clean = re.sub(r'^[一二三四五六七八九十]+[、\.\s]\s*', '', title_clean).strip()
        content_str = "；".join(content_list[:5])  # 限制前 5 列以防單格過長
        rows.append([title_clean, content_str])
        
    return rows

def is_6_box_grid(text_block):
    """安全檢核：確認文字代碼塊是否為標準 6 盒戰略畫布 (非一般網路拓撲或流程圖)"""
    lines = text_block.strip().split("\n")
    three_col_count = 0
    for line in lines:
        if line.count("│") >= 4:
            three_col_count += 1
    return three_col_count >= 5

def get_a3_canvas_data(client_dir):
    """
    自適應尋找並解析 A3 戰略畫布資料
    優先級：
    1. project_management_tables.md 內建的文字框格 (yuepin / henda 標準版)
    2. a3_strategy_canvas.md 內的文字框格
    3. a3_strategy_canvas.md 內部的標題大綱結構 (Fallback - dingtai 證券版)
    """
    pm_path = os.path.join(client_dir, "project_management_tables.md")
    if os.path.exists(pm_path):
        with open(pm_path, "r", encoding="utf-8") as f:
            content = f.read()
        match = re.search(r'```text\n(.*?)\n```', content, re.DOTALL)
        if match:
            text_block = match.group(1)
            if is_6_box_grid(text_block):
                rows = parse_a3_canvas_from_text(text_block)
                if rows:
                    return rows
                    
    a3_path = os.path.join(client_dir, "a3_strategy_canvas.md")
    if os.path.exists(a3_path):
        with open(a3_path, "r", encoding="utf-8") as f:
            content = f.read()
        match = re.search(r'```text\n(.*?)\n```', content, re.DOTALL)
        if match:
            text_block = match.group(1)
            if is_6_box_grid(text_block):
                rows = parse_a3_canvas_from_text(text_block)
                if rows:
                    return rows
        # Fallback to H2/H3 sections
        rows = parse_a3_from_sections(a3_path)
        if rows:
            return rows
            
    return []

def parse_markdown_tables(md_path):
    """
    解析 Markdown 檔案，提取標題與對應的表格數據
    回傳格式: [(sheet_title, headers, rows)]
    """
    if not os.path.exists(md_path):
        print(f"[錯誤] 找不到 Markdown 檔案: {md_path}")
        return []

    with open(md_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    tables = []
    current_sheet_title = "Sheet"
    in_table = False
    current_headers = []
    current_rows = []

    for line in lines:
        line_str = line.strip()
        
        # 1. 偵測 H2 或 H1 標題作為 Sheet 名稱的候選
        heading_match = re.match(r'^(##|#)\s+(.*)$', line_str)
        if heading_match:
            last_heading = heading_match.group(2).strip()
            # 排除 A3 戰略畫布或 Dashboard Panel，由後續專屬解析處理，避免衝突
            if "a3" in last_heading.lower() or "戰略畫布" in last_heading or "dashboard" in last_heading.lower() or "總表" in last_heading:
                in_table = False
                continue
            current_sheet_title = map_sheet_title(last_heading)
            continue

        # 2. 偵測表格行
        if line_str.startswith("|"):
            # 過濾掉對齊線 | :--- | :--- |
            if re.search(r'\|\s*:?-+:?\s*\|', line_str):
                continue
                
            cells = [clean_markdown_cell(c) for c in line_str.split("|")[1:-1]]
            
            if not in_table:
                # 進入新表格，第一行為 Header
                in_table = True
                current_headers = cells
                current_rows = []
            else:
                # 表格內容行
                current_rows.append(cells)
        else:
            if in_table:
                # 表格結束，保存起來
                tables.append((current_sheet_title, current_headers, current_rows))
                in_table = False

    # 若結尾仍在表格中
    if in_table:
        tables.append((current_sheet_title, current_headers, current_rows))

    # 3. 自動提取並增補 A3 Strategy Canvas 作為第 5 個工作表 (對齊 yuepin v2 機制)
    client_dir = os.path.dirname(md_path)
    a3_data = get_a3_canvas_data(client_dir)
    if a3_data:
        tables.append(("A3 Strategy Canvas", ["區塊", "內容"], a3_data))

    return tables

def export_tables_to_excel(tables, excel_out_path):
    """將解析出的表格寫入 Excel，並套用專業樣式"""
    wb = Workbook()
    # 刪除預設的 Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 樣式定義 (深藍色優雅表頭)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Microsoft JhengHei", size=10, color="000000")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for sheet_title, headers, rows in tables:
        # 新增工作表
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True  # 強制顯示網格線

        # 寫入表頭
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = cell_border
        
        # 寫入資料列
        for r_idx, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = cell_font
                cell.border = cell_border
                
                # 自動判斷靠左或靠中
                val_str = str(cell.value)
                if len(val_str) <= 10 or val_str.startswith("P0") or val_str.startswith("P1") or val_str.startswith("P2") or val_str.replace('.', '', 1).isdigit():
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # 設定自適應欄寬 (加上緩衝，最大支援到 65 寬度)
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # 中文字元計算雙倍長度
                val_len = sum(2 if ord(char) > 127 else 1 for char in val)
                if val_len > max_len:
                    max_len = val_len
            col_letter = get_column_letter(col[0].column)
            # 限制最大與最小寬度，避免過寬或過窄
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

        # 設定表頭列高為 28，資料列高為 24
        ws.row_dimensions[1].height = 28
        for r_idx in range(2, len(rows) + 2):
            ws.row_dimensions[r_idx].height = 24

    # 保存檔案
    os.makedirs(os.path.dirname(excel_out_path), exist_ok=True)
    wb.save(excel_out_path)
    print(f"[成功] 已成功將 Markdown 表格解析並儲存至實體 Excel: {excel_out_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Markdown 專案管理表轉 Excel 轉換器")
    parser.add_argument("--markdown", type=str, required=True, help="輸入的 Markdown 檔案路徑")
    parser.add_argument("--output", type=str, required=True, help="輸出的 Excel (.xlsx) 檔案路徑")
    args = parser.parse_args()

    tables = parse_markdown_tables(args.markdown)
    if tables:
        export_tables_to_excel(tables, args.output)
    else:
        print("[警告] 未能從檔案中解析到任何 Markdown 表格。")
