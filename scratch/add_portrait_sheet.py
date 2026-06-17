import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb_path = r'G:\我的雲端硬碟\AI_Talent\curriculum\unit_7_strategy\phoenix_ai_canvas_4plus1.xlsx'
wb = openpyxl.load_workbook(wb_path)

sheet_name = '附表-員工AI肖像授權'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 7) # Insert at index 7 (8th sheet)

# Styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(name='Microsoft JhengHei', bold=True, color='FFFFFF', size=14)
title_font = Font(name='Microsoft JhengHei', bold=True, size=18, color='1F4E78')
regular_font = Font(name='Microsoft JhengHei', size=12)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Set widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 25

# Title
ws.merge_cells('A1:D1')
ws['A1'] = '員工 AI 肖像永久商用授權同意書 (對照自診與修改草案)'
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:D2')
ws['A2'] = '評估貴公司目前的合約是否有涵蓋這三個保命條款，避免產生肖像侵權爭議。'
ws['A2'].font = Font(name='Microsoft JhengHei', size=12, italic=True)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Header Row
headers = ['條款類型', 'NG 合約寫法 (潛在風險)', '鳳凰 AI 黃金合約條款 (保命金句)', '貴公司現有合約是否具備？']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# Data Rows
data = [
    [
        '一、授權範圍與「AI 重製/生成權」雙重分離', 
        '「甲方同意無償授權乙方拍攝之照片及影片做為宣傳使用。」\n\n(風險：同意拍實體片，不代表同意被AI煉丹複製。若無註明AI數位孿生、生成式重製，未來AI生成的影片屬侵權。)', 
        '「授權人同意將包含其肖像之影音資料，授權被授權人進行包含但不限於『AI 模型訓練』、『特徵萃取』、『數位分身(Digital Twin)合成』、『語音克隆』等生成式 AI 重製行為，並同意被授權人享有由上述 AI 技術生成之衍生影音作品之完整著作財產權。」',
        ''
    ],
    [
        '二、明定「合理對價之買斷性質」',
        '「基於雙方僱傭關係，員工同意無償授權公司使用其肖像...」\n\n(風險：依據民法，無償贈與可依法撤銷。若無給付對價，員工未來隨時能以侵害人格權為由要求撤回授權。)',
        '「本授權為有償且不可撤回之授權。雙方同意授權對價為新台幣 ____ 元整 (或明定包含於每月薪資之特定加給中)。授權人確認已充分收受該對價，並承諾日後絕不以任何理由撤銷、終止本授權或主張任何額外之權利金。」',
        ''
    ],
    [
        '三、僱傭關係終止後之效力豁免',
        '「本同意書於雙方僱傭存續期間有效。」或未約定離職後效力。\n\n(風險：員工一旦離職，肖像授權自動失效，公司必須將所有含有該員工臉部特徵的 YouTube、官網行銷影片全數下架。)',
        '「本授權之效力不受雙方僱傭關係終止、解除、或變更之影響。授權人離職後，被授權人仍有權永久、不限地域、不限次數，將已生成及未來基於本合約授權生成之 AI 衍生影音內容，用於商業行銷、對外宣傳及所有合法之營運目的。」',
        ''
    ]
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = cell_value
        cell.font = regular_font
        cell.border = thin_border
        if col_idx == 1 or col_idx == 4:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
    ws.row_dimensions[row_idx].height = 100

# Data Validation for Column D
dv = DataValidation(type="list", formula1='"是 (已包含),否 (需修改),不適用"', allow_blank=True)
ws.add_data_validation(dv)
for row in range(4, 7):
    dv.add(ws[f'D{row}'])
    
# Add calculation at the bottom
ws.merge_cells('A8:C8')
ws['A8'] = 'AI 肖像安全度加權分數 (具備保命條款數 / 總條款數)'
ws['A8'].font = Font(name='Microsoft JhengHei', bold=True, size=14)
ws['A8'].alignment = Alignment(horizontal='right', vertical='center')
ws.row_dimensions[8].height = 25

ws['D8'] = '=COUNTIF(D4:D6, "是 (已包含)") & " / 3"'
ws['D8'].font = Font(name='Microsoft JhengHei', bold=True, size=14, color='FF0000')
ws['D8'].alignment = center_align

wb.save(wb_path)
print("Sheet '附表-員工AI肖像授權' created successfully!")
