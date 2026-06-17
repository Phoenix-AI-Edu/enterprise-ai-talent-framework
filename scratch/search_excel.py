# -*- coding: utf-8 -*-
import openpyxl
import os

def search_excel_content():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1.xlsx")
    out_path = os.path.join(base_dir, "scratch", "search_results.txt")
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found at: {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    with open(out_path, "w", encoding="utf-8") as f_out:
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            f_out.write(f"\n--- Checking Sheet: {sheetname} ---\n")
            for r_idx in range(1, sheet.max_row + 1):
                for c_idx in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r_idx, column=c_idx).value
                    if val and any(term in str(val) for term in ["RAG", "自診", "自評", "五維"]):
                        f_out.write(f"Row {r_idx}, Col {c_idx}: {val}\n")
    print(f"Done! Results written to: {out_path}")
                    
if __name__ == "__main__":
    search_excel_content()
