# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import os

def add_video_diagnostic_sheet():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1.xlsx")
    backup_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1_backup_video.xlsx")
    
    if not os.path.exists(file_path):
        print(f"Error: Excel file not found at {file_path}")
        return
        
    print("Creating backup...")
    try:
        shutil.copy2(file_path, backup_path)
        print(f"Backup created at: {backup_path}")
    except Exception as e:
        print(f"Failed to create backup: {e}")
        return
        
    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path)
    
    sheet_title = "附表-影片工具選型與財務自診"
    if sheet_title in wb.sheetnames:
        print(f"Sheet '{sheet_title}' already exists. Removing it to recreate...")
        wb.remove(wb[sheet_title])
        
    # Create sheet at index 6 (after '附表-RAG採購5維自診' which is index 5)
    print(f"Creating sheet '{sheet_title}'...")
    ws = wb.create_sheet(title=sheet_title, index=6)
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Calibri"
    
    # Style definitions
    title_font = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    title_fill = PatternFill(fill_type="solid", start_color="1B4F72", end_color="1B4F72") # Dark Blue
    title_align = Alignment(horizontal="center", vertical="center")
    
    desc_font = Font(name=font_family, size=9, bold=False, color="1A5276")
    desc_fill = PatternFill(fill_type="solid", start_color="D6EAF8", end_color="D6EAF8") # Light Blue
    desc_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="1B4F72", end_color="1B4F72") # Dark Blue
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_side = Side(border_style="thin", color="CCCCCC")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    normal_font = Font(name=font_family, size=10, bold=False, color="333333")
    bold_font = Font(name=font_family, size=10, bold=True, color="000000")
    
    # Row 1: Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "📋 附表：商用行銷影片 6 大工具選型與財務自診表 (AI Video Tools Selection & Financial Diagnostics)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = title_align
    ws.row_dimensions[1].height = 40
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = title_fill
        
    # Row 2: Description
    ws.merge_cells("A2:F2")
    ws["A2"] = "  ℹ️ 使用說明：本表供企業在採購或導入 AI 行銷影片生成工具 (如 Sora, Kling, Runway, HeyGen) 前進行評估。請針對以下 6 個維度進行自評打分 (每項 1~5 分)，以診斷工具的商用合規度、Token 財務 TCO 成本與特徵一致性就緒度。"
    ws["A2"].font = desc_font
    ws["A2"].fill = desc_fill
    ws["A2"].alignment = desc_align
    ws.row_dimensions[2].height = 35
    for col in range(1, 7):
        ws.cell(row=2, column=col).fill = desc_fill
        
    # Row 3: Headers
    headers = ["評估維度", "自評項目", "指標說明", "自評分數 (1-5)", "選型評估與財務指引", "備註 / 企業現況"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[3].height = 25
    
    # Content rows 4 to 9 (6 dimensions)
    video_dimensions = [
        [
            "維度一：肖像權合規風險",
            "肖像權合規風險\n(商用免責承諾)",
            "評估工具生成的虛擬人物或聲音是否會侵犯第三方肖像權或著作權，以及平台是否提供 B2B 客戶商用版權免責保護。",
            5,
            "1 分：開源或免費版，無任何授權保證，面臨高度侵權與客訴風險\n3 分：有付費套餐，但合約中無明確的智慧財產權商用免責保護條款\n5 分：企業級 SaaS，正式合約條款中明文承諾「商用安全免責保障」",
            ""
        ],
        [
            "維度二：TCO 算力財務控制",
            "Token 算力財務 TCO 控制\n(點數消耗與費用限制)",
            "影片生成算力消耗巨大，需要精算單次生成成本 (點數消耗率) 及企業整體的預算控制與防超付機制。",
            5,
            "1 分：單次生成成本極高，疊代修改費用昂貴，且缺乏預算限制功能\n3 分：提供基本套餐點數，但多版本修改容易超出額度，費用較高\n5 分：具備算力優化技術，單次消耗點數低，且具備部門限額管理",
            ""
        ],
        [
            "維度三：品牌聲調對齊度",
            "品牌聲調 DPO 對齊度\n(自建 Prompt 路由)",
            "影片整體風格、旁白配音與品牌調性的對齊程度。是否能接回內部 Prompt 路由器進行統一控管。",
            5,
            "1 分：僅能套用固定模板，配音生硬且無法對齊品牌形象調性\n3 分：支援自定義 Prompt 輸入，但風格無法精細對齊，仍需手動微調\n5 分：支援企業語音克隆與專屬 DPO 對齊，可無縫串接 Router Prompt",
            ""
        ],
        [
            "維度四：物理特徵一致性",
            "物理特徵一致性\n(LOGO與產品渲染精準度)",
            "評估影片中「產品細節、企業 LOGO、主角面部」在不同鏡頭切換間的一致性，防止影片產生扭曲變形等穿幫畫面。",
            5,
            "1 分：每秒畫面穿幫嚴重，LOGO 或產品細節在每幀快速變形扭曲\n3 分：主體畫面一致，但在大角度或快速運鏡時細節仍會失真穿幫\n5 分：支援 Reference Frame 或 3D 模型導入，特徵一致性達 98%",
            ""
        ],
        [
            "維度五：口型與語音克隆",
            "多國口型與語音克隆對齊\n(外銷口型同步)",
            "針對外銷多語系影片，評估 AI 生成的翻譯語音與畫面中人物口型的對齊精準度，避免不自然的視覺違和感。",
            5,
            "1 分：配音腔調極不自然且口型完全對不上，視覺違和感重\n3 分：支援多國配音，但口型同步率中等，有輕微滯後感\n5 分：具備高精度語音克隆 (Voice Cloning) 與 AI 口型對齊 (Lip-sync)",
            ""
        ],
        [
            "維度六：本地剪輯工具對接",
            "非侵入式本地剪輯對接\n(剪輯工作流與著作權)",
            "評估 AI 生成素材是否能與現有剪輯軟體 (Premiere/CapCut) 無縫整合，以及生成素材的版權歸屬保障。",
            5,
            "1 分：全網頁操作，無法與本地剪輯工作流連動，生成版權不明確\n3 分：可導出標準 MP4，但缺乏 API 支援，需手動下載與導入剪輯\n5 分：提供完整的 API 接口，支援本地自動化剪輯，素材產權歸屬企業",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(video_dimensions, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.font = normal_font
            
            # Alignments
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
        ws.row_dimensions[idx].height = 60
        
    # Row 10: Total Score & Diagnosis
    # Total Score Label
    cell_tot_lbl = ws.cell(row=10, column=3, value="🤖 總分自評加總：")
    cell_tot_lbl.font = bold_font
    cell_tot_lbl.alignment = Alignment(horizontal="right", vertical="center")
    cell_tot_lbl.border = cell_border
    
    # Total Score Formula
    cell_tot_val = ws.cell(row=10, column=4, value="=SUM(D4:D9)")
    cell_tot_val.font = Font(name=font_family, size=11, bold=True, color="FF5B35") # Brand Orange
    cell_tot_val.alignment = Alignment(horizontal="center", vertical="center")
    cell_tot_val.border = cell_border
    
    # Merge cells for diagnosis title and result
    ws.merge_cells("A10:B10")
    ws["A10"] = "📈 選型財務診斷結果："
    ws["A10"].font = bold_font
    ws["A10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A10"].border = cell_border
    
    ws.merge_cells("E10:F10")
    ws["E10"] = '=IF(D10<=15,"🔴 警訊：預算或技術一致性不足。不建議盲目採購高階 SaaS 專案，應先以小範圍免費工具驗證。",IF(D10<=23,"🟡 建議：採用 2/8 原則。使用 SaaS 進行草稿生成，並由一線剪輯師進行人機協作 (HITL) 覆核修剪。","🟢 優勢：就緒度高。可大規模採購 API 或企業級 SaaS，建立全自動行銷影片製作流水線。"))'
    ws["E10"].font = bold_font
    ws["E10"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["E10"].border = cell_border
    ws["E10"].fill = PatternFill(fill_type="solid", start_color="FCF3CF", end_color="FCF3CF") # Light gold fill
    
    ws.row_dimensions[10].height = 45
    
    # Adjust border for merged cells in openpyxl
    for col in [1, 2, 5, 6]:
        ws.cell(row=10, column=col).border = cell_border
        
    # Column Widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 30
    
    print("Saving workbook...")
    try:
        wb.save(file_path)
        print("Workbook saved successfully!")
    except PermissionError:
        print("\n[ERROR] Permission Denied! File might be open in Excel. Please close it.")
    except Exception as e:
        print(f"Error saving workbook: {e}")

if __name__ == "__main__":
    add_video_diagnostic_sheet()
