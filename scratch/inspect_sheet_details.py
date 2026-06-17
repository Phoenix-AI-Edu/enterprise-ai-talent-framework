# -*- coding: utf-8 -*-
import openpyxl
import os

def inspect_sheet_details():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1.xlsx")
    out_path = os.path.join(base_dir, "scratch", "sheet_details.txt")
    
    if not os.path.exists(excel_path):
        print(f"Excel file not found at: {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    with open(out_path, "w", encoding="utf-8") as f_out:
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            f_out.write(f"\n================ SHEET: {sheetname} ================\n")
            for r in range(1, 20):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 12)]
                f_out.write(f"Row {r:02d}: {row_vals}\n")
    print(f"Done! Results written to: {out_path}")

if __name__ == "__main__":
    inspect_sheet_details()
