# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import os
import sys

def add_rag_diagnostic_sheet():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1.xlsx")
    backup_path = os.path.join(base_dir, "curriculum", "unit_7_strategy", "phoenix_ai_canvas_4plus1_backup_rag.xlsx")
    
    if not os.path.exists(file_path):
        print(f"Error: Excel file not found at {file_path}")
        return
        
    print("Creating backup...")
    try:
        shutil.copy2(file_path, backup_path)
        print(f"Backup created at: {backup_path}")
    except Exception as e:
        print(f"Failed to create backup: {e}")
        return
        
    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path)
    
    sheet_title = "附表-RAG採購5維自診"
    if sheet_title in wb.sheetnames:
        print(f"Sheet '{sheet_title}' already exists. Removing it to recreate...")
        wb.remove(wb[sheet_title])
        
    # Create sheet at index 6 (after '主表三-試點與變革') or let's say index 5 (before '主表四-營運與治理')
    # Let's put it as sheet 6 (index 5)
    print(f"Creating sheet '{sheet_title}'...")
    ws = wb.create_sheet(title=sheet_title, index=5)
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Calibri"
    
    # Style definitions
    title_font = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    title_fill = PatternFill(fill_type="solid", start_color="1B4F72", end_color="1B4F72") # Dark Blue
    title_align = Alignment(horizontal="center", vertical="center")
    
    desc_font = Font(name=font_family, size=9, bold=False, color="1A5276")
    desc_fill = PatternFill(fill_type="solid", start_color="D6EAF8", end_color="D6EAF8") # Light Blue
    desc_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="1B4F72", end_color="1B4F72") # Dark Blue
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_side = Side(border_style="thin", color="CCCCCC")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    normal_font = Font(name=font_family, size=10, bold=False, color="333333")
    bold_font = Font(name=font_family, size=10, bold=True, color="000000")
    
    # Fill colors for dimensions
    dim_fill = PatternFill(fill_type="solid", start_color="F9EBEA", end_color="F9EBEA") # Very light red
    
    # Row 1: Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "📋 附表：企業 RAG 採購與數據成熟度 5 維自評表 (RAG Procurement & Data Readiness Diagnostics)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = title_align
    ws.row_dimensions[1].height = 40
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = title_fill
        
    # Row 2: Description
    ws.merge_cells("A2:F2")
    ws["A2"] = "  ℹ️ 使用說明：本表為 RAG 企業級知識庫採購前的自我成熟度評估表。請針對以下 5 個維度進行自評打分 (每項 1~5 分)，總分將會自動加總，並出具對應的顧問落地路徑診斷與避坑指南。"
    ws["A2"].font = desc_font
    ws["A2"].fill = desc_fill
    ws["A2"].alignment = desc_align
    ws.row_dimensions[2].height = 35
    for col in range(1, 7):
        ws.cell(row=2, column=col).fill = desc_fill
        
    # Row 3: Headers
    headers = ["評估維度", "自評項目", "指標說明", "自評分數 (1-5)", "成熟度評估指引", "備註 / 企業現況"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border
    ws.row_dimensions[3].height = 25
    
    # Content rows 4 to 8
    rag_dimensions = [
        [
            "維度一：檔案格式就緒度",
            "檔案格式就緒度\n(純文字、表格、掃描件%)",
            "評估企業內部資料庫、手冊與SOP文件的數位化就緒程度，防範資料讀不出而造成 RAG 系統失效。",
            5, # Default value, user can change
            "1 分：多為紙本掃描件或PDF圖片檔，缺乏OCR\n3 分：多為數位PDF，但含有複雜跨頁表格與排版雜訊\n5 分：全為結構化文字檔 (Markdown/TXT)，段落排版清晰",
            ""
        ],
        [
            "維度二：機密權限等級",
            "機密權限等級\n(公開/機密分流與AD/ACL)",
            "評估企業文件的安全層級分流，以及現有系統 (Active Directory/ACL) 是否具備串接權限管理能力。",
            5,
            "1 分：全為高度機密，且缺乏明確帳號角色權限控管\n3 分：已有部門分類，但無法與企業 AD/LDAP 權限連動\n5 分：公開/機密文件分流明確，且具備 API/AD 帳號對接",
            ""
        ],
        [
            "維度三：IT 本地維運能力",
            "IT 本地維運能力\n(Vector DB & MLOps)",
            "評估企業內部 IT 團隊的技術能力，判斷適合「API 買」還是「地端 Build」。",
            5,
            "1 分：僅有基礎系統維護人員，無 AI 或資料庫調優人員\n3 分：具備傳統 SQL 運維實力，但無 Vector DB 及 LLM 經驗\n5 分：有專業研發或 LLMOps 部門，可自主開發與微調模型",
            ""
        ],
        [
            "維度四：算力財務預算限制",
            "算力財務預算限制\n(SaaS訂閱 vs 地端GPU採購)",
            "評估可提撥的財務預算上限，決定算力架構與採購模式。",
            5,
            "1 分：預算極低，難以負擔 Token 呼叫或高額軟體合約費\n3 分：可負擔 SaaS 訂閱或 API 月費，但無法採購地端伺服器\n5 分：財務彈性大，可單獨編列預算採購 A100/H100 地端伺服器",
            ""
        ],
        [
            "維度五：業務場景容錯率",
            "業務場景容錯率\n(幻覺容忍度與CRAG修正)",
            "評估 AI 回答出錯時，企業與業務流程的容忍極限，決定是否需要高強度的防幻覺安全防線。",
            5,
            "1 分：零容忍，一旦幻覺將造成重大財務/法律責任 (如出帳、診斷)\n3 分：中度容忍，主要為內部輔助，回答必須標註引用出處頁碼\n5 分：高度容忍，用於腦力激盪、創意或行銷文案起草",
            ""
        ]
    ]
    
    for idx, row_data in enumerate(rag_dimensions, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.font = normal_font
            
            # Alignments
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
        ws.row_dimensions[idx].height = 60
        
    # Row 9: Total Score & Diagnosis
    # Total Score
    cell_tot_lbl = ws.cell(row=9, column=3, value="🤖 總分自評加總：")
    cell_tot_lbl.font = bold_font
    cell_tot_lbl.alignment = Alignment(horizontal="right", vertical="center")
    cell_tot_lbl.border = cell_border
    
    cell_tot_val = ws.cell(row=9, column=4, value="=SUM(D4:D8)")
    cell_tot_val.font = Font(name=font_family, size=11, bold=True, color="FF5B35") # Brand Orange
    cell_tot_val.alignment = Alignment(horizontal="center", vertical="center")
    cell_tot_val.border = cell_border
    
    # Merge cells for diagnosis title and result
    ws.merge_cells("A9:B9")
    ws["A9"] = "📈 成熟度診斷結果："
    ws["A9"].font = bold_font
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A9"].border = cell_border
    
    ws.merge_cells("E9:F9")
    ws["E9"] = '=IF(D9<=12,"🔴 警訊：基礎建設薄弱。建議首要做數據清洗，暫緩軟體採購，防範高額黑洞。",IF(D9<=19,"🟡 建議：採用 SaaS/API 混合架構，並配置嚴格的「真人雙簽 (HITL)」防線。","🟢 優勢：就緒度高。可自建或深度客製化 RAG 系統，直接開展試點。"))'
    ws["E9"].font = bold_font
    ws["E9"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["E9"].border = cell_border
    ws["E9"].fill = PatternFill(fill_type="solid", start_color="FCF3CF", end_color="FCF3CF") # Light gold fill
    
    ws.row_dimensions[9].height = 45
    
    # Adjust border for merged cells in openpyxl
    for col in [1, 2, 5, 6]:
        ws.cell(row=9, column=col).border = cell_border
        
    # Column Widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 30
    
    print("Saving workbook...")
    try:
        wb.save(file_path)
        print("Workbook saved successfully!")
    except PermissionError:
        print("\n[ERROR] Permission Denied! File might be open in Excel. Please close it.")
    except Exception as e:
        print(f"Error saving workbook: {e}")

if __name__ == "__main__":
    add_rag_diagnostic_sheet()
